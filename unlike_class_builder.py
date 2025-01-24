from abc import ABC, abstractmethod
import pandas as pd
import math
from openpyxl import Workbook
import date_handler
from openpyxl.styles import PatternFill
import shutil

# CREATES AN UNLIKE REQUEST CLASS
# account = the client account that the request is based on
# cusips = a list of cusips being pulled from an unlike pdf or other location
# holdings = an exported holdings sheet from the sponsors portal. This is stored as a filepath
# initials = The builder
class UnlikeSellRequest:
    def __init__(self, account, cusips, holdings, initials, count):
        self.account = account
        self.cusips = cusips
        self.holdings = holdings
        self.initials = initials
        self.count = count

    @abstractmethod
    def build_uploader(self):
        return


# Because each sponsor has a different method for unlikes (different holdings tables, uploads, etc),
# we need different methods for building the unlike QC components
class MSUnlike(UnlikeSellRequest):

    def __init__(self, account, cusips, holdings, initials, count, mgr_code):
        super().__init__(account, cusips, holdings, initials, count)
        # initialize it as blank?
        self.mgr_code = mgr_code



    def build_uploader(self):
        # upload sheet templates
        ms_upload_template_sheet = 'MS_upload_blank.csv'
        ms_upload_df = pd.read_csv(ms_upload_template_sheet)
        ms_holdings_cusips_df = pd.read_excel(self.holdings, header=7)
        ms_holdings_full_info_df = pd.read_excel(self.holdings, header=None, nrows=6)

        ofc_num = self.account[:3]
        dean_num = self.account[3:]

        ms_holdings_cusips_df['ToSell'] = ms_holdings_cusips_df['Symbol/CUSIP'].apply \
            (lambda x: True if x in self.cusips or len(x) <= 6 else False)

        ms_holdings_sells_df = ms_holdings_cusips_df.query('ToSell == True')

        # formatting the upload sheet
        ms_upload_df['SYMBOL'] = ms_holdings_sells_df['Symbol/CUSIP']
        ms_upload_df['QUANTITY'] = ms_holdings_sells_df['Position']
        ms_upload_df['OFC #'] = ofc_num
        ms_upload_df['ACCT #'] = dean_num
        ms_upload_df['BS'] = 'S'
        ms_upload_df['MGR CODE'] = self.mgr_code

        # condition to round stocks down. WORK IN PROGRESS but works for now
        # mutual funds have X at the end. These can have frac-shares (decimal points in the qty)
        # regular stocks MUST BE ROUNDED DOWN ON UPLOAD SHEET.
        # ETF's also rounded down like stocks.
        round_condition = ~ms_upload_df['SYMBOL'].str.endswith('X')

        ms_upload_df.loc[round_condition, 'QUANTITY'] = ms_upload_df.loc[round_condition, 'QUANTITY'].apply(math.floor)
        print(ms_upload_df.head())

        return [ms_holdings_sells_df, ms_upload_df]

        # create the MS upload template
        # ms_upload_df.to_csv(
        #     f'G:\\ManagedAccounts\\Operations\\Corporate Ladders\\TradeWorkflowAutomation\\Settlements\\unlikes\\output\\{self.initials}MS {date_handler.today_crunched}-{self.count}.csv',
        #     index=0)

    def build_sell_reference(self):

        ms_holdings_cusips_df = pd.read_excel(self.holdings, header=7)
        ms_holdings_main_info_df = pd.read_excel(self.holdings, header=None, nrows=6)

        # create the holdings template. With highlighted cusips
        highlight = PatternFill(start_color='FFF000', end_color='FFF000', fill_type='solid')

        wb = Workbook()
        ws = wb.active

        output_loc = (
            f'G:\\ManagedAccounts\\Operations\\Corporate Ladders\\TradeWorkflowAutomation\\Settlements\\unlikes\\'
            f'output\\{self.account} MS Holdings {date_handler.today_standard}.xlsx')

        for r_idx, row in ms_holdings_main_info_df.iterrows():
            for c_idx, value in enumerate(row):
                ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)

        for c_idx, col_name in enumerate(ms_holdings_cusips_df):
            ws.cell(row=len(ms_holdings_main_info_df) + 2, column=c_idx + 1, value=col_name)

        for r_idx, row in ms_holdings_cusips_df.iterrows():
            for c_idx, value in enumerate(row):
                cell = ws.cell(row=r_idx + len(ms_holdings_main_info_df) + 3, column=c_idx + 1, value=value)

                if c_idx == 7 and value:
                    print('highlighting sell')
                    for col in range(1, len(row) + 1):
                        ws.cell(row=r_idx + len(ms_holdings_main_info_df) + 3, column=col).fill = highlight

        return wb

        # copy the original holdings file to the output folder

        # og_output_loc = (
        #     f'G:\\ManagedAccounts\\Operations\\Corporate Ladders\\TradeWorkflowAutomation\\Settlements\\unlikes\\'
        #     f'output')
        # new_file_name = f'reference {self.account} MS Holdings {date_handler.today_standard}.xlsx'
        # copy_and_rename(self.holdings, og_output_loc, new_file_name)


class MLUnlike(UnlikeSellRequest):
    def __init__(self, account, cusips, holdings, initials, count):
        super().__init__(account, cusips, holdings, initials, count)



    def build_uploader(self):
        return


def copy_and_rename(src_path, dest_path, new_name):
    # Copy the file
    shutil.copy(src_path, dest_path)

    # Rename the copied file
    # new_path = f"{dest_path}/{new_name}"
    # shutil.move(f"{dest_path}/{src_path}", new_path)


def good_to_sell(cusip, listofcusips):
    # check if item is on the unlike list
    # if the item is a stock, sell.
    if cusip in listofcusips or len(cusip) <= 6:
        return 'Sell'
    else:
        return ''
